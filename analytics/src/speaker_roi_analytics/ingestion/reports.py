"""Error-report rendering, and the one place this package writes tabular files.

Why this module exists
----------------------
plan.md §10.2 requires every upload attempt to produce "machine-readable JSON
plus human-readable CSV error reports", and plan.md §10.3 requires the original
row number to survive into them so the uploader can open the file they sent and
go to the exact line.  Those reports are the product's entire feedback loop for
data quality: if they are wrong, vague, or unsafe, the platform silently trains
its users to ignore them.

Three properties are non-negotiable here.

**Safety on write.** Everything in a report is derived from an untrusted upload
(plan.md §15 "Treat all uploaded data as untrusted") — most obviously the
uploader's own column headings, which we echo back so they can find the column
we are complaining about.  A cell whose text begins with ``=``, ``+``, ``-``,
``@``, a tab or a carriage return is executed as a formula by Excel, LibreOffice
and Google Sheets when the report is opened.  That turns "here are your errors"
into remote code execution against the analyst who opened it, with the payload
supplied by whoever produced the upload.  :func:`neutralise_formula` is applied
to **every** cell this module writes, in CSV and in XLSX alike, and every writer
in this package (including :mod:`.templates`) routes through it.

**Separation of outcomes.** A file-level failure and a row-level failure are
different events and are reported in different places: ``file_error`` /
``mapping`` issues carry no row number and mean *nothing was loaded*, while row
issues name a row and mean *that row alone* was rejected or quarantined.  The
CSV keeps them distinguishable by leaving ``row_number`` blank for the former
and by carrying an explicit ``disposition`` column for the latter, so a reader
can never mistake "37 bad rows" for "the file was refused".

**Determinism.** Reports are checksummed and diffed in tests and in CI, so
nothing here reads the clock: ``generated_at`` is injected by the caller.
"""

from __future__ import annotations

import csv
import datetime as dt
import json
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from decimal import Decimal
from enum import Enum
from pathlib import Path
from typing import Any, Final

from speaker_roi_core.enums import IssueSeverity

from .issues import Issue
from .validate import RowDisposition, ValidationOutcome

__all__ = [
    "FORMULA_TRIGGERS",
    "NEUTRALISER",
    "REPORT_COLUMNS",
    "REPORT_SCHEMA_VERSION",
    "ReportPaths",
    "build_json_report",
    "neutralise_formula",
    "report_rows",
    "write_csv",
    "write_csv_report",
    "write_json_report",
    "write_reports",
    "write_xlsx_report",
]

#: The leading characters a spreadsheet treats as "this cell is code".
#:
#: ``=`` and ``@`` open a formula outright; ``+`` and ``-`` do too (``-1+1`` is a
#: formula, not a number); ``\t`` and ``\r`` are included because Excel strips
#: leading whitespace before deciding, so ``"\t=cmd|'/c calc'!A0"`` reaches the
#: formula parser with the tab discarded.  This is the OWASP CSV-injection set
#: and matches plan.md §15.
FORMULA_TRIGGERS: Final[frozenset[str]] = frozenset({"=", "+", "-", "@", "\t", "\r"})

#: Prefix that demotes a cell to text.  A leading apostrophe is the portable
#: signal across Excel, LibreOffice and Sheets; it is visible in the cell, which
#: is the point — a reader should be able to see that we defused something.
NEUTRALISER: Final[str] = "'"

#: Bumped whenever the JSON report's shape changes in a way a consumer must
#: notice.  The API and the portal both branch on it.
REPORT_SCHEMA_VERSION: Final[str] = "1.0"

#: Column order of the human-readable CSV report.  Locators first, because the
#: first question is always "where", and remediation last, because it is the
#: longest.
REPORT_COLUMNS: Final[tuple[str, ...]] = (
    "row_number",
    "sheet",
    "column",
    "field",
    "severity",
    "scope",
    "disposition",
    "code",
    "title",
    "message",
    "remediation",
    "gate",
    "category",
    "docs_anchor",
)


# ---------------------------------------------------------------------------
# Formula-injection neutralisation (plan.md §15)
# ---------------------------------------------------------------------------


def neutralise_formula(value: object) -> str:
    """Render ``value`` as text that no spreadsheet will execute.

    Applied to every cell written by this package.  The rule is deliberately
    blunt — *any* leading trigger character gets the prefix, including on a
    plain negative number such as ``-3``, which renders as ``'-3``.

    That over-reach is a considered trade-off.  Distinguishing ``-3`` (a
    number) from ``-3+cmd()`` (a payload) means parsing the cell, and a parser
    is exactly the kind of component whose edge cases become the bypass.  The
    cost of being blunt is a cosmetic apostrophe in an error report and in the
    sample row of a template; the cost of being clever is code execution on an
    analyst's workstation.  We take the apostrophe.

    Embedded newlines and carriage returns are also collapsed to spaces: a
    header containing ``\\n`` would otherwise split one report row into two and
    misalign every column after it, which is a cheap way to hide a finding.
    """
    text = _cell_text(value)
    if not text:
        return text
    text = text.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    if text[0] in FORMULA_TRIGGERS:
        return NEUTRALISER + text
    return text


def _cell_text(value: object) -> str:
    """Canonical text for a value, before neutralisation.

    Booleans render lowercase and dates ISO-8601 so a report round-trips
    through the same coercion rules that produced it.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date | dt.time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, Enum):
        return str(value.value)
    return str(value)


def _safe_row(values: Iterable[object]) -> list[str]:
    return [neutralise_formula(value) for value in values]


# ---------------------------------------------------------------------------
# Tabular rendering
# ---------------------------------------------------------------------------


def report_rows(outcome: ValidationOutcome) -> tuple[tuple[str, ...], ...]:
    """The error report as rows of already-neutralised text, in report order.

    Ordering is file-level findings first (``row_number`` blank, sorted by code)
    then row findings in file order — the order someone fixing the file works
    in, because a header fault invalidates every row beneath it and must be
    read first.
    """
    dispositions = {row.row_number: row.disposition for row in outcome.rows}
    issues: list[Issue] = list(outcome.issues)
    if outcome.file_error is not None and outcome.file_error not in issues:
        issues.insert(0, outcome.file_error)
    ordered = sorted(issues, key=lambda issue: issue.sort_key())
    return tuple(tuple(_report_row(issue, dispositions)) for issue in ordered)


def _report_row(
    issue: Issue,
    dispositions: Mapping[int, RowDisposition],
) -> list[str]:
    definition = issue.definition
    if issue.row_number is None:
        scope = "FILE"
        disposition = ""
    else:
        scope = "ROW"
        disposition = str(dispositions.get(issue.row_number, ""))
    return _safe_row(
        (
            "" if issue.row_number is None else issue.row_number,
            issue.sheet or "",
            issue.column or "",
            issue.field or "",
            issue.severity.value,
            scope,
            disposition,
            issue.code.value,
            definition.title,
            issue.message,
            definition.remediation,
            definition.gate.value,
            definition.category.value,
            definition.docs_anchor,
        )
    )


def build_json_report(
    outcome: ValidationOutcome,
    *,
    generated_at: dt.datetime | None = None,
    upload_id: str | None = None,
) -> dict[str, Any]:
    """The machine-readable report (plan.md §10.2).

    ``generated_at`` is a parameter and not a clock read so two runs over the
    same bytes produce byte-identical JSON; the API passes the request
    timestamp, tests pass a fixed one.
    """
    payload = outcome.as_dict()
    return {
        "report_schema_version": REPORT_SCHEMA_VERSION,
        "generated_at": None if generated_at is None else generated_at.isoformat(),
        "upload_id": upload_id,
        "verdict": _verdict(outcome),
        **payload,
        "issue_counts_by_code": _counts_by_code(outcome),
    }


def _verdict(outcome: ValidationOutcome) -> str:
    """One word for "what happened", so a UI does not re-derive the rules.

    The four states are distinct on purpose (the coordinator's rule that
    quarantine and error must not be collapsed): ``REFUSED`` means the file was
    never opened for data, ``BLOCKED`` means it parsed but a file-scope gate
    failed, ``PARTIAL`` means some rows are loadable and others are not, and
    ``ACCEPTED`` means everything loaded.
    """
    if outcome.is_file_level_failure:
        return "REFUSED"
    if outcome.has_blocking_errors:
        return "BLOCKED"
    summary = outcome.summary
    unusable = summary.rejected + summary.quarantined
    if summary.loadable_rows == 0:
        return "BLOCKED"
    return "PARTIAL" if unusable else "ACCEPTED"


def _counts_by_code(outcome: ValidationOutcome) -> dict[str, int]:
    counts: dict[str, int] = {}
    for issue in outcome.issues:
        counts[issue.code.value] = counts.get(issue.code.value, 0) + 1
    return dict(sorted(counts.items()))


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class ReportPaths:
    """What :func:`write_reports` produced. ``None`` means "not requested"."""

    json_path: Path | None = None
    csv_path: Path | None = None
    xlsx_path: Path | None = None

    def existing(self) -> tuple[Path, ...]:
        return tuple(p for p in (self.json_path, self.csv_path, self.xlsx_path) if p)


def write_csv(
    path: Path,
    header: Sequence[str],
    rows: Iterable[Sequence[object]],
    *,
    neutralise: bool = True,
) -> Path:
    """Write a CSV that is safe to open in a spreadsheet.

    ``newline=""`` is required by :mod:`csv` on Windows or every record gains a
    stray blank line.  ``utf-8-sig`` is deliberate: Excel on Windows assumes the
    system codepage for a BOM-less UTF-8 file and mangles every non-ASCII name
    in the report, and a mangled report is one the user cannot act on.

    ``neutralise=False`` exists only for callers writing values they generated
    themselves and have already checked; it is never used for upload-derived
    content.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(_safe_row(header) if neutralise else list(header))
        for row in rows:
            writer.writerow(_safe_row(row) if neutralise else [_cell_text(v) for v in row])
    return path


def write_csv_report(outcome: ValidationOutcome, path: Path) -> Path:
    """The human-readable CSV error report (plan.md §10.2)."""
    return write_csv(path, REPORT_COLUMNS, report_rows(outcome), neutralise=False)


def write_json_report(
    outcome: ValidationOutcome,
    path: Path,
    *,
    generated_at: dt.datetime | None = None,
    upload_id: str | None = None,
) -> Path:
    """The machine-readable JSON error report (plan.md §10.2).

    JSON needs no formula neutralisation — nothing executes a JSON string — but
    the messages it carries were already redacted at construction time by
    :func:`~speaker_roi_analytics.ingestion.issues.make_issue`.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = build_json_report(outcome, generated_at=generated_at, upload_id=upload_id)
    path.write_text(
        json.dumps(payload, indent=2, sort_keys=False, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return path


#: Fill colours for the severity column of the XLSX report. Muted on purpose:
#: an error workbook that looks like a warning label gets closed, not read.
_SEVERITY_FILL: Final[Mapping[str, str]] = {
    IssueSeverity.ERROR.value: "FFF4E5E5",
    IssueSeverity.QUARANTINE.value: "FFFDF3E0",
    IssueSeverity.WARNING.value: "FFFBF8E3",
    IssueSeverity.INFO.value: "FFEFF3F8",
}

_REPORT_WIDTHS: Final[Mapping[str, int]] = {
    "row_number": 11,
    "sheet": 14,
    "column": 24,
    "field": 24,
    "severity": 12,
    "scope": 8,
    "disposition": 22,
    "code": 34,
    "title": 40,
    "message": 70,
    "remediation": 70,
    "gate": 26,
    "category": 14,
    "docs_anchor": 34,
}


def write_xlsx_report(outcome: ValidationOutcome, path: Path) -> Path:
    """An optional error *workbook*: the same rows, filterable and frozen.

    Same neutralisation as the CSV.  openpyxl decides a cell is a formula from
    a leading ``=`` in the string it is handed, so writing a raw header such as
    ``=SUM(A:A)`` back into a report would store a live formula in the file we
    hand the analyst — the neutralised text is what reaches the cell.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    rows = report_rows(outcome)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    bold = Font(bold=True)
    for index, (key, value) in enumerate(_summary_pairs(outcome), start=1):
        summary_sheet.cell(row=index, column=1, value=neutralise_formula(key)).font = bold
        summary_sheet.cell(row=index, column=2, value=neutralise_formula(value))
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 60

    sheet = workbook.create_sheet("Issues")
    header_fill = PatternFill("solid", fgColor="FFE8EDF3")
    for column, name in enumerate(REPORT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column, value=name)
        cell.font = bold
        cell.fill = header_fill
        sheet.column_dimensions[get_column_letter(column)].width = _REPORT_WIDTHS.get(name, 18)
    wrap = Alignment(vertical="top", wrap_text=True)
    severity_index = REPORT_COLUMNS.index("severity")
    for row_index, values in enumerate(rows, start=2):
        fill_colour = _SEVERITY_FILL.get(values[severity_index])
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column, value=value)
            cell.alignment = wrap
            if fill_colour:
                cell.fill = PatternFill("solid", fgColor=fill_colour)
    sheet.freeze_panes = "A2"
    if rows:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(REPORT_COLUMNS))}{len(rows) + 1}"

    workbook.save(path)
    workbook.close()
    return path


def _summary_pairs(outcome: ValidationOutcome) -> tuple[tuple[str, str], ...]:
    summary = outcome.summary
    return (
        ("Dataset", str(outcome.contract.dataset_type)),
        ("Contract version", outcome.contract.version),
        ("File", "" if outcome.path is None else outcome.path.name),
        ("Verdict", _verdict(outcome)),
        ("Rows read", str(summary.total_rows)),
        ("Rows loadable", str(summary.loadable_rows)),
        ("Rows quarantined", str(summary.quarantined)),
        ("Rows rejected", str(summary.rejected)),
        ("Rows superseded", str(summary.superseded)),
        ("Errors", str(summary.error_count)),
        ("Quarantines", str(summary.quarantine_count)),
        ("Warnings", str(summary.warning_count)),
        ("Issue list truncated", "true" if summary.issues_truncated else "false"),
    )


def write_reports(
    outcome: ValidationOutcome,
    out_dir: Path,
    *,
    stem: str = "validation_report",
    generated_at: dt.datetime | None = None,
    upload_id: str | None = None,
    include_xlsx: bool = False,
) -> ReportPaths:
    """Write the JSON and CSV reports (and optionally the workbook) together.

    Both formats always, because they serve different readers: the JSON is what
    the API stores and the portal renders, the CSV is what a data steward opens
    next to the file they are fixing.  plan.md §10.2 asks for both by name.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = write_json_report(
        outcome,
        out_dir / f"{stem}.json",
        generated_at=generated_at,
        upload_id=upload_id,
    )
    csv_path = write_csv_report(outcome, out_dir / f"{stem}.csv")
    xlsx_path = write_xlsx_report(outcome, out_dir / f"{stem}.xlsx") if include_xlsx else None
    return ReportPaths(json_path=json_path, csv_path=csv_path, xlsx_path=xlsx_path)
