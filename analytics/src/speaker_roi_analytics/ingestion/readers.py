"""Streaming CSV and XLSX readers that never lose a row number.

Everything downstream of this module addresses problems by their **original**
1-based position in the uploaded file — physical CSV line, or spreadsheet row
(plan.md §10.3: "Preserve original row number in all validation errors").  An
error report that says "row 4,812" is actionable; one that says "record 4,730
after we skipped some blanks" is not, and a reviewer who cannot find the cell
stops trusting the report.  That single requirement drives most of the design
choices here:

* CSV is parsed with the standard library's :mod:`csv` module rather than a
  columnar reader, because a record containing an embedded newline inside a
  quoted field occupies several physical lines and only a record-at-a-time
  parser can report where it actually started.  ``csv.reader.line_num`` counts
  physical lines, so the start line of every record is exact.
* XLSX is read with ``openpyxl`` in ``read_only`` mode, and the row number comes
  from the cell's own ``.row`` attribute rather than from an enumeration
  counter, so a sheet with blank rows in the middle still reports true
  spreadsheet coordinates.

Both readers stream in chunks regardless of file size (plan.md §10.3: "Stream
parse; never load the entire file into memory"), enforce the byte and row
ceilings from PLAN_REVIEW F-12, and refuse a set of file shapes outright:

``.xlsm`` / ``.xls`` / ``.xlsb``
    Macro-enabled and legacy binary workbooks are refused by extension *and* by
    content inspection, because renaming ``book.xlsm`` to ``book.xlsx`` must not
    get a VBA project into the ingestion pipeline.

Encrypted workbooks
    An encrypted OOXML file is an OLE2 compound document, not a zip. It is
    detected by magic bytes and refused with its own code, so the uploader is
    told to remove the password rather than being shown "corrupt file".

Zip bombs and absurd sheet dimensions
    A 40 KB workbook that expands to 8 GB, or a sheet claiming a million
    columns, is rejected before openpyxl is asked to parse it.

XLSX is always read with ``data_only=True``: the platform consumes the values
Excel last cached, and never evaluates a formula. Evaluating uploaded formulas
would be both a correctness problem (we would disagree with what the uploader
saw) and an attack surface.
"""

from __future__ import annotations

import codecs
import csv
import datetime as dt
import zipfile
from collections.abc import Iterator, Sequence
from dataclasses import dataclass, field
from pathlib import Path
from types import TracebackType
from typing import IO, Any, Final

import chardet
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from speaker_roi_analytics.ingestion.issues import Issue, IssueCode, make_issue
from speaker_roi_core.enums import FileFormat

__all__ = [
    "CANDIDATE_DELIMITERS",
    "DEFAULT_LIMITS",
    "MAX_LIMITS",
    "CsvRowSource",
    "ReadPlan",
    "ReaderError",
    "ReaderLimits",
    "RowSource",
    "SourceRow",
    "XlsxRowSource",
    "detect_format",
    "open_row_source",
    "sniff_csv",
]


# ---------------------------------------------------------------------------
# limits
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class ReaderLimits:
    """Ceilings applied while reading, independent of contract rules.

    PLAN_REVIEW F-12 fixes the configured maximum at 250 MB / 1,000,000 rows and
    the shipped default much lower.  The low default is not timidity: a 25 MB
    demo tenant that rejects a 200 MB accidental full-history export gives the
    uploader an immediate, comprehensible answer, whereas accepting it produces a
    twenty-minute wait and a memory profile nobody budgeted for.  Parsing is
    chunked either way, so raising the limit changes what is *allowed*, never
    how much memory is used.
    """

    max_bytes: int = 25 * 1024 * 1024
    max_rows: int = 200_000
    #: Rows handed to the validator at a time. Bounds peak memory to roughly
    #: ``chunk_size`` rows regardless of file size.
    chunk_size: int = 5_000
    max_columns: int = 512
    #: A single cell larger than this is a pasted document, not a data value.
    max_field_bytes: int = 131_072
    #: Uncompressed-to-compressed ratio above which a workbook is treated as a
    #: zip bomb. Legitimate spreadsheets compress well but not absurdly.
    max_compression_ratio: int = 200
    #: Bytes sampled for encoding and delimiter detection.
    sniff_bytes: int = 64 * 1024
    #: chardet confidence below which we refuse to trust the guess.
    encoding_confidence_floor: float = 0.65

    def with_(self, **overrides: Any) -> ReaderLimits:
        """Return a copy with ``overrides`` applied — limits are per-tenant config."""
        current = {
            "max_bytes": self.max_bytes,
            "max_rows": self.max_rows,
            "chunk_size": self.chunk_size,
            "max_columns": self.max_columns,
            "max_field_bytes": self.max_field_bytes,
            "max_compression_ratio": self.max_compression_ratio,
            "sniff_bytes": self.sniff_bytes,
            "encoding_confidence_floor": self.encoding_confidence_floor,
        }
        current.update(overrides)
        return ReaderLimits(**current)


#: Shipped default (PLAN_REVIEW F-12 "demo default").
DEFAULT_LIMITS: Final[ReaderLimits] = ReaderLimits()

#: The configured ceiling. A deployment may raise ``DEFAULT_LIMITS`` up to here
#: but not past it: beyond this size the right answer is a database extract, not
#: a spreadsheet upload.
MAX_LIMITS: Final[ReaderLimits] = ReaderLimits(max_bytes=262_144_000, max_rows=1_000_000)

#: Delimiters the sniffer will consider. Deliberately short: every additional
#: candidate makes a wrong guess more likely, and these four cover essentially
#: every real export.
CANDIDATE_DELIMITERS: Final[tuple[str, ...]] = (",", ";", "\t", "|")

_DELIMITER_NAMES: Final[dict[str, str]] = {
    ",": "comma",
    ";": "semicolon",
    "\t": "tab",
    "|": "pipe",
}

_ZIP_MAGIC: Final[bytes] = b"PK\x03\x04"
_OLE2_MAGIC: Final[bytes] = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
_UTF8_BOM: Final[bytes] = codecs.BOM_UTF8
_UTF16_BOMS: Final[tuple[bytes, ...]] = (codecs.BOM_UTF16_LE, codecs.BOM_UTF16_BE)

_MACRO_MEMBER: Final[str] = "xl/vbaProject.bin"
_ENCRYPTED_MEMBER_HINTS: Final[tuple[str, ...]] = ("EncryptedPackage", "EncryptionInfo")


class ReaderError(Exception):
    """A file-level refusal, carrying the :class:`Issue` to show the uploader.

    File-level problems are exceptions rather than accumulated findings because
    there is nothing to accumulate: if the workbook is encrypted there are no
    rows to report on.  Row-level problems never raise.
    """

    def __init__(self, issue: Issue) -> None:
        super().__init__(issue.message)
        self.issue = issue


def _fail(code: IssueCode, **params: object) -> ReaderError:
    return ReaderError(make_issue(code, **params))


def _raise_csv_field_limit(target: int) -> None:
    """Raise :func:`csv.field_size_limit` to ``target`` if it is currently lower.

    Only ever raised, never lowered: the limit is process-global, and a reader
    configured for large cells must not shrink the ceiling for a concurrently
    running one.
    """
    current = csv.field_size_limit()
    if target > current:
        csv.field_size_limit(target)


# ---------------------------------------------------------------------------
# read plan
# ---------------------------------------------------------------------------


@dataclass(frozen=True, slots=True)
class SourceRow:
    """One physical record, still as text.

    ``row_number`` is where the record *starts* in the file, 1-based.  For a CSV
    record containing an embedded newline this is the first of the several
    physical lines it occupies, which is where a person looking at the file
    would put their cursor.
    """

    row_number: int
    values: tuple[str, ...]
    sheet: str | None = None


@dataclass(frozen=True, slots=True)
class ReadPlan:
    """What the reader concluded about the file, before committing to it.

    This is returned to the caller *for confirmation*.  plan.md §10.3 asks for
    delimiter detection with confirmation rather than a silent guess, because
    reading a semicolon file as comma-delimited does not fail — it produces one
    enormous column, a completely different set of validation errors, and a
    support conversation that starts three steps away from the real problem.
    """

    file_format: FileFormat
    encoding: str
    encoding_confidence: float
    has_bom: bool
    delimiter: str
    delimiter_confidence: float
    quotechar: str = '"'
    delimiter_confirmed: bool = False
    encoding_confirmed: bool = False
    sheet_name: str | None = None
    sheet_names: tuple[str, ...] = ()
    hidden_sheet_names: tuple[str, ...] = ()
    byte_size: int = 0
    issues: tuple[Issue, ...] = ()

    @property
    def delimiter_label(self) -> str:
        return _DELIMITER_NAMES.get(self.delimiter, repr(self.delimiter))

    def as_dict(self) -> dict[str, Any]:
        """Shown in the upload wizard's confirmation step."""
        return {
            "file_format": self.file_format.value,
            "encoding": self.encoding,
            "encoding_confidence": round(self.encoding_confidence, 3),
            "has_bom": self.has_bom,
            "delimiter": self.delimiter_label,
            "delimiter_confidence": round(self.delimiter_confidence, 3),
            "delimiter_confirmed": self.delimiter_confirmed,
            "sheet_name": self.sheet_name,
            "sheet_names": list(self.sheet_names),
            "hidden_sheet_names": list(self.hidden_sheet_names),
            "byte_size": self.byte_size,
            "issues": [i.as_dict() for i in self.issues],
        }


# ---------------------------------------------------------------------------
# format detection and shared pre-flight
# ---------------------------------------------------------------------------


def detect_format(path: Path) -> FileFormat:
    """Map an extension onto a supported format, refusing the rest by name.

    Extension is checked before content so the refusal message can name what the
    uploader actually did ("this is a macro-enabled workbook") instead of
    describing a parse failure.  Content is checked *as well*, in
    :func:`_inspect_workbook_bytes` — a rename must not be a bypass.
    """
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return FileFormat.CSV
    if suffix in {".txt", ".tsv"}:
        # Same parser; the delimiter sniffer sorts out which separator it uses.
        return FileFormat.CSV
    if suffix == ".xlsx":
        return FileFormat.XLSX
    if suffix == ".jsonl":
        return FileFormat.JSONL
    if suffix == ".xlsm":
        raise _fail(IssueCode.FILE_MACRO_ENABLED_WORKBOOK, extension=suffix)
    if suffix in {".xls", ".xlsb"}:
        raise _fail(IssueCode.FILE_LEGACY_BINARY_WORKBOOK, extension=suffix)
    raise _fail(
        IssueCode.FILE_UNSUPPORTED_EXTENSION,
        extension=suffix or "(none)",
        allowed=".csv, .tsv, .txt, .xlsx",
    )


def _preflight(path: Path, limits: ReaderLimits) -> int:
    """Existence, emptiness and size, before a single byte is parsed."""
    if not path.is_file():
        raise _fail(IssueCode.FILE_NOT_FOUND, name=path.name)
    size = path.stat().st_size
    if size == 0:
        raise _fail(IssueCode.FILE_EMPTY, name=path.name)
    if size > limits.max_bytes:
        raise _fail(
            IssueCode.FILE_TOO_LARGE,
            size=f"{size / 1_048_576:.1f} MB",
            limit=f"{limits.max_bytes / 1_048_576:.0f} MB",
        )
    return size


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


def _detect_encoding(sample: bytes, limits: ReaderLimits) -> tuple[str, float, bool, Issue | None]:
    """BOM first, chardet second, an explicit documented fallback third.

    A BOM is definitive and cheap, so it wins outright.  Without one, chardet is
    a statistical guess: below the confidence floor we do **not** quietly accept
    it, because mis-decoding turns a valid file into a wall of mojibake errors
    that look like the uploader's fault.  We fall back to UTF-8 and say so.
    """
    if sample.startswith(_UTF8_BOM):
        return "utf-8-sig", 1.0, True, None
    for bom in _UTF16_BOMS:
        if sample.startswith(bom):
            return "utf-16", 1.0, True, None

    guess = chardet.detect(sample)
    encoding = (guess.get("encoding") or "").lower()
    confidence = float(guess.get("confidence") or 0.0)

    if not encoding:
        return (
            "utf-8",
            0.0,
            False,
            make_issue(IssueCode.FILE_ENCODING_LOW_CONFIDENCE, encoding="utf-8", confidence="0.00"),
        )

    # chardet reports plain "ascii" for files that are ASCII so far; UTF-8 is a
    # strict superset, so decoding as UTF-8 is always at least as correct and
    # tolerates a non-ASCII byte appearing later in the file.
    if encoding in {"ascii", "utf-8"}:
        return "utf-8", max(confidence, 0.99), False, None

    if confidence < limits.encoding_confidence_floor:
        return (
            "utf-8",
            confidence,
            False,
            make_issue(
                IssueCode.FILE_ENCODING_LOW_CONFIDENCE,
                encoding=encoding,
                confidence=f"{confidence:.2f}",
            ),
        )
    return encoding, confidence, False, None


def _score_delimiter(lines: Sequence[str], delimiter: str) -> tuple[int, float]:
    """Field count and consistency for one candidate delimiter.

    Consistency matters more than raw count: a comma-delimited file read with
    semicolons yields one field on every line (perfectly consistent, useless),
    while the true delimiter yields the *same* count greater than one on every
    line.  Scoring on "same count, and that count > 1" separates the two without
    needing to understand the data.
    """
    counts: list[int] = []
    for line in lines:
        in_quotes = False
        count = 1
        index = 0
        while index < len(line):
            char = line[index]
            if char == '"':
                if in_quotes and index + 1 < len(line) and line[index + 1] == '"':
                    index += 2
                    continue
                in_quotes = not in_quotes
            elif char == delimiter and not in_quotes:
                count += 1
            index += 1
        counts.append(count)
    if not counts:
        return 0, 0.0
    modal = max(set(counts), key=counts.count)
    if modal <= 1:
        return 0, 0.0
    consistency = counts.count(modal) / len(counts)
    return modal, consistency


def sniff_csv(
    path: Path,
    *,
    limits: ReaderLimits = DEFAULT_LIMITS,
    encoding: str | None = None,
    delimiter: str | None = None,
) -> ReadPlan:
    """Work out how to read a delimited file, and say how sure we are.

    An explicitly supplied ``encoding`` or ``delimiter`` is treated as confirmed
    and short-circuits the corresponding detection.  Otherwise the guess is
    returned with a confidence and, when two candidates are close, a
    :data:`IssueCode.FILE_DELIMITER_AMBIGUOUS` warning.  If nothing scores at
    all the file is refused with
    :data:`IssueCode.FILE_DELIMITER_UNCONFIRMED` rather than parsed as a
    single-column file — plan.md §10.3 wants a confirmation step, not a
    plausible-looking wrong answer.
    """
    size = _preflight(path, limits)
    with path.open("rb") as handle:
        sample_bytes = handle.read(limits.sniff_bytes)

    issues: list[Issue] = []
    if encoding is not None:
        resolved_encoding, confidence, has_bom = encoding, 1.0, sample_bytes.startswith(_UTF8_BOM)
    else:
        resolved_encoding, confidence, has_bom, encoding_issue = _detect_encoding(
            sample_bytes, limits
        )
        if encoding_issue is not None:
            issues.append(encoding_issue)

    try:
        text = sample_bytes.decode(resolved_encoding, errors="replace")
    except LookupError as exc:  # unknown codec name from chardet
        raise _fail(IssueCode.FILE_ENCODING_UNDETECTED, encoding=resolved_encoding) from exc

    lines = [line for line in text.splitlines()[:50] if line.strip()]
    if not lines:
        raise _fail(IssueCode.FILE_EMPTY, name=path.name)

    if delimiter is not None:
        return ReadPlan(
            file_format=FileFormat.CSV,
            encoding=resolved_encoding,
            encoding_confidence=confidence,
            has_bom=has_bom,
            delimiter=delimiter,
            delimiter_confidence=1.0,
            delimiter_confirmed=True,
            encoding_confirmed=encoding is not None,
            byte_size=size,
            issues=tuple(issues),
        )

    scores = {candidate: _score_delimiter(lines, candidate) for candidate in CANDIDATE_DELIMITERS}
    ranked = sorted(
        scores.items(),
        key=lambda item: (item[1][1], item[1][0], -CANDIDATE_DELIMITERS.index(item[0])),
        reverse=True,
    )
    best, (best_fields, best_consistency) = ranked[0]
    if best_fields == 0:
        raise _fail(
            IssueCode.FILE_DELIMITER_UNCONFIRMED,
            candidates=", ".join(_DELIMITER_NAMES[d] for d in CANDIDATE_DELIMITERS),
        )

    runner_up, (runner_fields, runner_consistency) = ranked[1]
    if runner_fields > 0 and runner_consistency >= best_consistency:
        issues.append(
            make_issue(
                IssueCode.FILE_DELIMITER_AMBIGUOUS,
                first=_DELIMITER_NAMES[best],
                second=_DELIMITER_NAMES[runner_up],
            )
        )

    return ReadPlan(
        file_format=FileFormat.CSV,
        encoding=resolved_encoding,
        encoding_confidence=confidence,
        has_bom=has_bom,
        delimiter=best,
        delimiter_confidence=best_consistency,
        delimiter_confirmed=False,
        encoding_confirmed=encoding is not None,
        byte_size=size,
        issues=tuple(issues),
    )


# ---------------------------------------------------------------------------
# XLSX pre-flight
# ---------------------------------------------------------------------------


def _inspect_workbook_bytes(path: Path, limits: ReaderLimits) -> None:
    """Refuse dangerous or impossible workbooks before openpyxl touches them.

    Checks, in order of how badly we want to avoid the next step:

    1. **OLE2 magic** — an encrypted OOXML package, or a genuine ``.xls``, both
       of which are compound documents. Refused by their own codes so the
       message can tell the uploader to remove the password.
    2. **Not a zip** — a renamed CSV or a truncated download.
    3. **Macro project inside** — ``xl/vbaProject.bin`` present regardless of the
       extension. Renaming does not launder a macro workbook.
    4. **Zip bomb** — total uncompressed size, and the compression ratio, both
       bounded. ``zipfile`` reports declared sizes without extracting, so this
       costs nothing and prevents an out-of-memory abort during parsing.
    """
    with path.open("rb") as handle:
        header = handle.read(8)

    if header.startswith(_OLE2_MAGIC):
        # Distinguish "encrypted .xlsx" from "genuine legacy .xls" by extension,
        # since both are OLE2 containers and the remedy is different.
        if path.suffix.lower() == ".xlsx":
            raise _fail(IssueCode.FILE_ENCRYPTED_WORKBOOK, name=path.name)
        raise _fail(IssueCode.FILE_LEGACY_BINARY_WORKBOOK, extension=path.suffix.lower())

    if not header.startswith(_ZIP_MAGIC):
        raise _fail(IssueCode.FILE_CORRUPT, name=path.name, reason="not a valid .xlsx package")

    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
            infos = archive.infolist()
    except zipfile.BadZipFile as exc:
        raise _fail(IssueCode.FILE_CORRUPT, name=path.name, reason="damaged zip container") from exc

    if any(name.startswith(_ENCRYPTED_MEMBER_HINTS) for name in names):
        raise _fail(IssueCode.FILE_ENCRYPTED_WORKBOOK, name=path.name)
    if _MACRO_MEMBER in names:
        raise _fail(IssueCode.FILE_MACRO_ENABLED_WORKBOOK, extension=path.suffix.lower())

    total_uncompressed = sum(info.file_size for info in infos)
    total_compressed = sum(info.compress_size for info in infos) or 1
    ratio = total_uncompressed / total_compressed
    ceiling = limits.max_bytes * limits.max_compression_ratio
    if ratio > limits.max_compression_ratio or total_uncompressed > ceiling:
        raise _fail(
            IssueCode.FILE_COMPRESSION_RATIO_SUSPICIOUS,
            ratio=f"{ratio:.0f}:1",
            limit=f"{limits.max_compression_ratio}:1",
        )


def _select_worksheet(
    workbook: Any,
    *,
    sheet: str | None,
    allow_hidden: bool,
) -> tuple[Worksheet, tuple[str, ...], tuple[str, ...]]:
    """Pick the sheet to read, defaulting to the first *visible* one.

    Hidden sheets are skipped by default because they are overwhelmingly
    scratch calculations, pivot caches and old versions.  Silently reading one
    would produce a confidently wrong load, so selecting a hidden sheet has to
    be an explicit act (``allow_hidden``) or an explicit name.
    """
    all_names = tuple(workbook.sheetnames)
    hidden = tuple(name for name in all_names if workbook[name].sheet_state != "visible")

    if sheet is not None:
        if sheet not in all_names:
            raise _fail(
                IssueCode.FILE_SHEET_NOT_FOUND,
                sheet=sheet,
                available=", ".join(all_names[:10]),
            )
        return workbook[sheet], all_names, hidden

    for name in all_names:
        worksheet = workbook[name]
        if allow_hidden or worksheet.sheet_state == "visible":
            return worksheet, all_names, hidden

    raise _fail(IssueCode.FILE_NO_VISIBLE_SHEET, name=", ".join(all_names[:10]))


# ---------------------------------------------------------------------------
# row sources
# ---------------------------------------------------------------------------


@dataclass
class _SourceState:
    header: tuple[str, ...] = ()
    header_row_number: int = 0
    rows_yielded: int = 0
    issues: list[Issue] = field(default_factory=list)


class RowSource:
    """Base class: a header plus a chunked stream of :class:`SourceRow`.

    Subclasses are context managers so the underlying handle or workbook is
    always released — a read-only openpyxl workbook holds an open zip, and
    leaking those across a long-running validation service exhausts file
    descriptors well before it exhausts memory.
    """

    plan: ReadPlan

    def __init__(self, plan: ReadPlan, limits: ReaderLimits) -> None:
        self.plan = plan
        self._limits = limits
        self._state = _SourceState()

    # -- context manager --------------------------------------------------
    def __enter__(self) -> RowSource:
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc: BaseException | None,
        tb: TracebackType | None,
    ) -> None:
        self.close()

    def close(self) -> None:  # pragma: no cover - overridden
        return None

    # -- interface --------------------------------------------------------
    @property
    def header(self) -> tuple[str, ...]:
        return self._state.header

    @property
    def header_row_number(self) -> int:
        return self._state.header_row_number

    @property
    def issues(self) -> tuple[Issue, ...]:
        """File-level findings gathered while reading (never row-level)."""
        return (*self.plan.issues, *self._state.issues)

    def iter_chunks(self) -> Iterator[tuple[SourceRow, ...]]:  # pragma: no cover - overridden
        raise NotImplementedError

    def iter_rows(self) -> Iterator[SourceRow]:
        for chunk in self.iter_chunks():
            yield from chunk

    # -- shared guards ----------------------------------------------------
    def _check_row_budget(self) -> None:
        self._state.rows_yielded += 1
        if self._state.rows_yielded > self._limits.max_rows:
            raise _fail(
                IssueCode.FILE_TOO_MANY_ROWS,
                rows=str(self._state.rows_yielded),
                limit=str(self._limits.max_rows),
            )

    def _check_width(self, width: int, row_number: int) -> None:
        if width > self._limits.max_columns:
            raise _fail(
                IssueCode.FILE_TOO_MANY_COLUMNS,
                columns=str(width),
                limit=str(self._limits.max_columns),
                row=str(row_number),
            )


class CsvRowSource(RowSource):
    """Record-at-a-time CSV reader that reports true physical line numbers."""

    def __init__(self, path: Path, plan: ReadPlan, limits: ReaderLimits) -> None:
        super().__init__(plan, limits)
        self._path = path
        # csv enforces a global per-field ceiling and raises a bare csv.Error
        # when it is hit. Lifting it just past our own limit means the oversized
        # cell is reported as FILE_FIELD_TOO_LARGE with a row number, not as an
        # opaque parse failure.
        _raise_csv_field_limit(limits.max_field_bytes + 2)
        # newline="" is required by the csv module: it must see the raw line
        # endings so a quoted field containing \r\n survives intact.
        self._handle: IO[str] = path.open("r", encoding=plan.encoding, newline="", errors="strict")
        self._reader = csv.reader(self._handle, delimiter=plan.delimiter, quotechar=plan.quotechar)
        self._read_header()

    def close(self) -> None:
        if not self._handle.closed:
            self._handle.close()

    def _read_header(self) -> None:
        """Advance to the first non-blank record and treat it as the header."""
        previous_line = 0
        try:
            for record in self._reader:
                start_line = previous_line + 1
                previous_line = self._reader.line_num
                if any(cell.strip() for cell in record):
                    self._state.header = tuple(cell.strip() for cell in record)
                    self._state.header_row_number = start_line
                    self._check_width(len(record), start_line)
                    return
        except UnicodeDecodeError as exc:
            raise _fail(
                IssueCode.FILE_DECODE_ERROR, encoding=self.plan.encoding, position=str(exc.start)
            ) from exc
        except csv.Error as exc:
            raise _fail(IssueCode.FILE_CORRUPT, name=self._path.name, reason=str(exc)) from exc
        raise _fail(IssueCode.SCHEMA_NO_HEADER_ROW, name=self._path.name)

    def iter_chunks(self) -> Iterator[tuple[SourceRow, ...]]:
        chunk: list[SourceRow] = []
        previous_line = self._reader.line_num
        try:
            for record in self._reader:
                start_line = previous_line + 1
                previous_line = self._reader.line_num
                if not any(cell.strip() for cell in record):
                    # A trailing newline, or a blank separator line inside the
                    # file. Skipped silently but *not* renumbered: the next real
                    # row keeps its true physical line number.
                    continue
                self._check_row_budget()
                self._check_width(len(record), start_line)
                self._check_field_sizes(record, start_line)
                chunk.append(SourceRow(row_number=start_line, values=tuple(record)))
                if len(chunk) >= self._limits.chunk_size:
                    yield tuple(chunk)
                    chunk = []
        except UnicodeDecodeError as exc:
            raise _fail(
                IssueCode.FILE_DECODE_ERROR, encoding=self.plan.encoding, position=str(exc.start)
            ) from exc
        except csv.Error as exc:
            raise _fail(IssueCode.FILE_CORRUPT, name=self._path.name, reason=str(exc)) from exc
        if chunk:
            yield tuple(chunk)

    def _check_field_sizes(self, record: Sequence[str], row_number: int) -> None:
        for cell in record:
            if len(cell) > self._limits.max_field_bytes:
                raise _fail(
                    IssueCode.FILE_FIELD_TOO_LARGE,
                    row=str(row_number),
                    limit=str(self._limits.max_field_bytes),
                )


class XlsxRowSource(RowSource):
    """Read-only, values-only worksheet reader preserving spreadsheet row numbers."""

    def __init__(
        self,
        path: Path,
        plan: ReadPlan,
        limits: ReaderLimits,
        *,
        sheet: str | None,
        allow_hidden_sheets: bool,
    ) -> None:
        super().__init__(plan, limits)
        self._path = path
        try:
            # read_only: stream rows instead of building a full object graph.
            # data_only: use the values Excel cached. Formulas are never
            # evaluated - the platform reports what the uploader saw, and an
            # uploaded formula is not something we want to execute.
            self._workbook = load_workbook(
                filename=str(path), read_only=True, data_only=True, keep_links=False
            )
        except zipfile.BadZipFile as exc:
            raise _fail(IssueCode.FILE_CORRUPT, name=path.name, reason="damaged workbook") from exc
        except (KeyError, ValueError, TypeError) as exc:
            raise _fail(IssueCode.FILE_CORRUPT, name=path.name, reason=type(exc).__name__) from exc

        worksheet, all_names, hidden = _select_worksheet(
            self._workbook, sheet=sheet, allow_hidden=allow_hidden_sheets
        )
        self._worksheet = worksheet
        self.plan = ReadPlan(
            file_format=plan.file_format,
            encoding=plan.encoding,
            encoding_confidence=plan.encoding_confidence,
            has_bom=plan.has_bom,
            delimiter=plan.delimiter,
            delimiter_confidence=plan.delimiter_confidence,
            delimiter_confirmed=True,
            encoding_confirmed=True,
            sheet_name=worksheet.title,
            sheet_names=all_names,
            hidden_sheet_names=hidden,
            byte_size=plan.byte_size,
            issues=plan.issues,
        )
        self._check_dimensions()
        self._rows = self._worksheet.iter_rows(values_only=False)
        self._read_header()

    def close(self) -> None:
        self._workbook.close()

    def _check_dimensions(self) -> None:
        """Reject a sheet whose declared extent alone exceeds the limits.

        ``max_row``/``max_column`` come from the sheet's declared dimension, so
        this is a cheap guard against a workbook that claims 1,048,576 rows
        because someone formatted an entire column.
        """
        max_row = self._worksheet.max_row or 0
        max_column = self._worksheet.max_column or 0
        if max_column > self._limits.max_columns:
            raise _fail(
                IssueCode.FILE_SHEET_DIMENSIONS_EXCEEDED,
                sheet=self._worksheet.title,
                rows=str(max_row),
                columns=str(max_column),
            )
        if max_row > self._limits.max_rows:
            raise _fail(
                IssueCode.FILE_SHEET_DIMENSIONS_EXCEEDED,
                sheet=self._worksheet.title,
                rows=str(max_row),
                columns=str(max_column),
            )

    @staticmethod
    def _row_number(cells: Sequence[Any], fallback: int) -> int:
        for cell in cells:
            row = getattr(cell, "row", None)
            if isinstance(row, int):
                return row
        return fallback

    @staticmethod
    def _text(cell: Any) -> str:
        """Render a cached cell value as text without inventing precision.

        openpyxl hands back native Python types.  ``float`` is stringified via
        ``repr`` semantics that would turn ``1.0`` into ``"1.0"`` — harmless —
        but an integral float becomes ``"30.0"``, which then fails an integer
        field for no good reason.  Integral floats are therefore rendered
        without the fractional part, and dates are rendered ISO so the
        coercion layer's date parsing sees its canonical form.
        """
        value = getattr(cell, "value", None)
        if value is None:
            return ""
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, dt.datetime):
            # Excel stores every date as a datetime. A midnight timestamp is a
            # date the user typed as a date, and rendering it "2026-03-01" keeps
            # it readable in the error report instead of "2026-03-01T00:00:00".
            if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
                return value.date().isoformat()
            return value.isoformat()
        if isinstance(value, dt.date | dt.time):
            return value.isoformat()
        return str(value)

    def _read_header(self) -> None:
        for counter, cells in enumerate(self._rows, start=1):
            row_number = self._row_number(cells, counter)
            values = tuple(self._text(cell) for cell in cells)
            if any(v.strip() for v in values):
                self._state.header = tuple(v.strip() for v in values)
                self._state.header_row_number = row_number
                self._check_width(len(values), row_number)
                return
        raise _fail(IssueCode.SCHEMA_NO_HEADER_ROW, name=self._path.name)

    def iter_chunks(self) -> Iterator[tuple[SourceRow, ...]]:
        chunk: list[SourceRow] = []
        counter = self._state.header_row_number
        sheet_name = self._worksheet.title
        for cells in self._rows:
            counter += 1
            row_number = self._row_number(cells, counter)
            counter = row_number
            values = tuple(self._text(cell) for cell in cells)
            if not any(v.strip() for v in values):
                continue
            self._check_row_budget()
            self._check_width(len(values), row_number)
            chunk.append(SourceRow(row_number=row_number, values=values, sheet=sheet_name))
            if len(chunk) >= self._limits.chunk_size:
                yield tuple(chunk)
                chunk = []
        if chunk:
            yield tuple(chunk)


# ---------------------------------------------------------------------------
# entry point
# ---------------------------------------------------------------------------


def open_row_source(
    path: Path | str,
    *,
    limits: ReaderLimits = DEFAULT_LIMITS,
    encoding: str | None = None,
    delimiter: str | None = None,
    sheet: str | None = None,
    allow_hidden_sheets: bool = False,
    require_delimiter_confirmation: bool = False,
) -> RowSource:
    """Open a file for streaming, refusing anything the platform will not read.

    ``require_delimiter_confirmation`` is what the interactive upload wizard
    sets: it turns an unconfirmed delimiter guess into a
    :data:`IssueCode.FILE_DELIMITER_UNCONFIRMED` refusal so the user is asked
    before ten thousand rows are parsed the wrong way.  Batch callers that have
    already agreed a format with the supplier pass ``delimiter=`` instead, which
    counts as confirmation.
    """
    resolved = Path(path)
    file_format = detect_format(resolved)
    size = _preflight(resolved, limits)

    if file_format is FileFormat.JSONL:
        raise _fail(
            IssueCode.FILE_UNSUPPORTED_EXTENSION,
            extension=".jsonl",
            allowed=".csv, .tsv, .txt, .xlsx",
        )

    if file_format is FileFormat.CSV:
        plan = sniff_csv(resolved, limits=limits, encoding=encoding, delimiter=delimiter)
        if require_delimiter_confirmation and not plan.delimiter_confirmed:
            raise _fail(
                IssueCode.FILE_DELIMITER_UNCONFIRMED,
                candidates=f"detected {plan.delimiter_label}",
            )
        return CsvRowSource(resolved, plan, limits)

    _inspect_workbook_bytes(resolved, limits)
    plan = ReadPlan(
        file_format=FileFormat.XLSX,
        encoding="utf-8",
        encoding_confidence=1.0,
        has_bom=False,
        delimiter=",",
        delimiter_confidence=1.0,
        delimiter_confirmed=True,
        encoding_confirmed=True,
        byte_size=size,
    )
    return XlsxRowSource(
        resolved,
        plan,
        limits,
        sheet=sheet,
        allow_hidden_sheets=allow_hidden_sheets,
    )
